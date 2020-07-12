import openpyxl

path = r"C:\Users\PawelGilecki\Desktop\SeleniumTest\data1.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook["Arkusz1"]

rows = sheet.max_row
columns = sheet.max_column

print(rows)
print(columns)

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r,column=c).value,end="    ")
    print()